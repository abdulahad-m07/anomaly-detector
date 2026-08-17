"""Tests for the automatic layout detection (auto mode)."""

import os
import sys
import tempfile
import unittest
from datetime import datetime, timedelta

from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.auto_detect import (
    AutoDetector,
    detect_date_format,
    guess_metric_profile,
    to_number,
)
from modules.data_reader import DataReader


def write_workbook(path, sheet_name, header, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)


class TestAutoDetect(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def _detect(self, path):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            return AutoDetector().detect(wb)
        finally:
            wb.close()

    def test_detects_long_format_with_odd_headers(self):
        path = os.path.join(self.tmp, "long.xlsx")
        start = datetime(2026, 8, 1)
        rows = []
        for d in range(10):
            ts = (start + timedelta(days=d)).strftime("%d/%m/%Y %H:%M")
            rows.append([ts, "Apples", 100 + d])
            rows.append([ts, "Bananas", 50 + d])
        write_workbook(path, "Data", ["Timestamp", "Indicator", "Amount"], rows)

        detection = self._detect(path)
        self.assertEqual(detection.layout, "long")
        self.assertEqual(detection.date_col, 0)
        self.assertEqual(detection.metric_col, 1)
        self.assertEqual(detection.value_col, 2)

    def test_detects_wide_format(self):
        path = os.path.join(self.tmp, "wide.xlsx")
        start = datetime(2026, 8, 1)
        rows = [
            [(start + timedelta(days=d)).strftime("%Y-%m-%d"), 100 + d, 50 + d, 3]
            for d in range(10)
        ]
        write_workbook(path, "Fruit Store",
                       ["Date", "Apples", "Bananas", "Website Crashes"], rows)

        detection = self._detect(path)
        self.assertEqual(detection.layout, "wide")
        self.assertEqual(detection.metric_names,
                         ["Apples", "Bananas", "Website Crashes"])

    def test_reader_reads_wide_file(self):
        path = os.path.join(self.tmp, "wide.xlsx")
        start = datetime(2026, 8, 1)
        rows = [
            [(start + timedelta(days=d)).strftime("%Y-%m-%d"), 100 + d, 50 + d, 3]
            for d in range(10)
        ]
        write_workbook(path, "Fruit Store",
                       ["Date", "Apples", "Bananas", "Website Crashes"], rows)

        records = DataReader({"file_path": path, "auto": True}).read_records()
        names = {r["metric_name"] for r in records}
        self.assertIn("Apples", names)
        self.assertIn("Bananas", names)
        self.assertIn("Website Crashes", names)
        self.assertEqual(len(records), 30)

    def test_reader_auto_reads_long_file(self):
        path = os.path.join(self.tmp, "long.xlsx")
        start = datetime(2026, 8, 1)
        rows = []
        for d in range(5):
            ts = (start + timedelta(days=d)).strftime("%d/%m/%Y %H:%M")
            rows.append([ts, "Apples", 100 + d])
            rows.append([ts, "Bananas", 50 + d])
        write_workbook(path, "Data", ["Timestamp", "Indicator", "Amount"], rows)

        records = DataReader({"file_path": path, "auto": True}).read_records()
        self.assertEqual(len(records), 10)
        self.assertEqual(records[0]["metric_name"], "Apples")

    def test_to_number_handles_currency(self):
        self.assertEqual(to_number("$1,234.56"), 1234.56)
        self.assertEqual(to_number("45%"), 45.0)
        self.assertEqual(to_number(42), 42.0)
        with self.assertRaises(ValueError):
            to_number("N/A")

    def test_date_format_detection(self):
        fmt = detect_date_format(["01/08/2026 10:00", "02/08/2026 11:30"])
        self.assertEqual(fmt, "%d/%m/%Y %H:%M")

    def test_metric_profile_guess(self):
        self.assertEqual(guess_metric_profile("Sales Revenue")["unit"], "USD")
        self.assertEqual(guess_metric_profile("Server CPU")["priority"], "HIGH")
        self.assertEqual(guess_metric_profile("Random Thing")["priority"],
                         "MEDIUM")


if __name__ == "__main__":
    unittest.main()
