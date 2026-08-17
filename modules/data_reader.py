"""Data ingestion module.

Reads an Excel file (.xlsx), picks the worksheet, and cleans every row
(dates -> datetime, values -> float, missing data handled per policy)
into a list of record dicts.

Two modes:

- **Configured** (``auto: false``): uses the column names given in config.
- **Automatic** (``auto: true``): inspects the workbook and detects the
  sheet, the date column, and whether the data is long format
  (Date | Metric | Value) or wide format (Date | MetricA | MetricB ...).
"""

import os
import shutil
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from modules.auto_detect import AutoDetector, to_number


class DataReaderError(Exception):
    """Raised when the Excel file cannot be read or parsed."""


# Fallback date formats tried when the configured format does not match.
_FALLBACK_FORMATS = [
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%d/%m/%Y",
]


class DataReader:
    """Reads and cleans Excel files into a list of record dicts."""

    def __init__(self, excel_config, logger=None):
        self.config = excel_config
        self.logger = logger

    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------
    def _log(self, message, level="info"):
        """Write a log line to the logger if one was given."""
        if self.logger is not None:
            getattr(self.logger, level)(message)
        else:
            print(f"[{level.upper()}] {message}")

    def _pick_worksheet(self, workbook):
        """Return the configured worksheet, defaulting to the first sheet."""
        worksheet_name = self.config.get("worksheet_name")
        if worksheet_name:
            if worksheet_name not in workbook.sheetnames:
                raise DataReaderError(
                    f"Worksheet '{worksheet_name}' not found. "
                    f"Available: {workbook.sheetnames}"
                )
            return workbook[worksheet_name]
        return workbook.worksheets[0]

    def _map_columns(self, header):
        """Locate the index of each required column in the header row.

        Returns a dict {date, metric, value} -> column index.
        """
        config_keys = {
            "date": self.config.get("date_column", "Date"),
            "metric": self.config.get("metric_column", "Metric Name"),
            "value": self.config.get("value_column", "Value"),
        }
        header = [str(h).strip() if h is not None else "" for h in header]

        mapping = {}
        for key, column_name in config_keys.items():
            try:
                mapping[key] = header.index(column_name)
            except ValueError:
                raise DataReaderError(
                    f"Required column '{column_name}' not found in "
                    f"worksheet header. Header is: {header}"
                ) from None
        return mapping

    def _parse_timestamp(self, raw, fmt_override=None):
        """Turn a raw cell into a datetime object.

        fmt_override (optional) is a date format the auto-detector already
        confirmed, tried before the fallback list.
        """
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw
        if isinstance(raw, (int, float)):
            # Excel stores dates as a serial number (days since 1899-12-30).
            serial = float(raw)
            if serial >= 60:
                serial -= 1  # Excel's famous 1900 leap-year bug
            return datetime(1899, 12, 30) + timedelta(days=serial)
        text = str(raw).strip()
        fmt = self.config.get("datetime_format")
        candidates = []
        if fmt_override and fmt_override != fmt:
            candidates.append(fmt_override)
        if fmt:
            candidates.append(fmt)
        candidates += _FALLBACK_FORMATS
        for candidate in candidates:
            if candidate is None:
                continue
            try:
                return datetime.strptime(text, candidate)
            except (ValueError, TypeError):
                continue
        self._log(f"Could not parse timestamp '{raw}' - skipping row", "warning")
        return None

    def _parse_value(self, raw):
        """Turn a raw cell into a float, or None if it is not numeric."""
        if raw is None:
            return None
        try:
            return to_number(raw)
        except ValueError:
            self._log(
                f"Non-numeric value '{raw}' encountered - skipping row",
                "warning",
            )
            return None

    def _iter_row_pairs(self, worksheet):
        """Yield (header, raw_row) for every row in the worksheet."""
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise DataReaderError("Excel file is empty (no header row).")
        return header, rows

    # ------------------------------------------------------------------
    # configured mode
    # ------------------------------------------------------------------
    def _read_configured(self, worksheet):
        """Read using column names from config (long format assumed)."""
        header, rows = self._iter_row_pairs(worksheet)
        mapping = self._map_columns(header)
        records = []
        last_value = {}  # used by the "interpolate" missing-value policy
        policy = self.config.get("missing_value_policy", "skip")

        for raw_row in rows:
            ncols = max(mapping.values()) + 1
            row = list(raw_row) + [None] * (ncols - len(raw_row))
            timestamp = self._parse_timestamp(row[mapping["date"]])
            if timestamp is None:
                continue
            metric_name = str(row[mapping["metric"]] or "").strip()
            if not metric_name:
                continue
            value = self._parse_value(row[mapping["value"]])
            if value is None:
                if policy == "interpolate" and metric_name in last_value:
                    value = last_value[metric_name]
                else:
                    continue
            else:
                last_value[metric_name] = value
            records.append(
                {
                    "timestamp": timestamp,
                    "metric_name": metric_name,
                    "metric_value": value,
                }
            )
        return records

    # ------------------------------------------------------------------
    # automatic mode
    # ------------------------------------------------------------------
    def _read_auto(self, worksheet, detection):
        """Read using the detected layout (long or wide)."""
        header, rows = self._iter_row_pairs(worksheet)
        records = []
        last_value = {}
        policy = self.config.get("missing_value_policy", "skip")
        fmt = detection.date_format

        if detection.layout == "wide":
            metric_cols = detection.numeric_cols
            for raw_row in rows:
                row = list(raw_row)
                timestamp = self._parse_timestamp(
                    row[detection.date_col] if detection.date_col is not None
                    else None, fmt
                )
                if timestamp is None:
                    continue
                for col in metric_cols:
                    value = self._parse_value(
                        row[col] if col < len(row) else None
                    )
                    metric_name = header[col]
                    if value is None:
                        if policy == "interpolate" and metric_name in last_value:
                            value = last_value[metric_name]
                        else:
                            continue
                    else:
                        last_value[metric_name] = value
                    records.append(
                        {
                            "timestamp": timestamp,
                            "metric_name": metric_name,
                            "metric_value": value,
                        }
                    )
            return records

        # long format
        for raw_row in rows:
            row = list(raw_row)
            timestamp = self._parse_timestamp(
                row[detection.date_col] if detection.date_col is not None
                else None, fmt
            )
            if timestamp is None:
                continue
            metric_name = str(
                row[detection.metric_col] if detection.metric_col is not None
                else ""
            ).strip()
            if not metric_name:
                continue
            value = self._parse_value(
                row[detection.value_col] if detection.value_col is not None
                else None
            )
            if value is None:
                if policy == "interpolate" and metric_name in last_value:
                    value = last_value[metric_name]
                else:
                    continue
            else:
                last_value[metric_name] = value
            records.append(
                {
                    "timestamp": timestamp,
                    "metric_name": metric_name,
                    "metric_value": value,
                }
            )
        return records

    # ------------------------------------------------------------------
    # main entry point
    # ------------------------------------------------------------------
    def read_records(self):
        """Read the Excel file and return a list of record dicts.

        Each record looks like::

            {
                "timestamp": datetime(2026, 8, 16, 10, 30),
                "metric_name": "Sales Revenue",
                "metric_value": 45000.5,
            }
        """
        file_path = self.config.get("file_path")
        if not file_path or not os.path.exists(file_path):
            raise DataReaderError(f"Excel file not found: {file_path}")

        # Copy the file to a temp location first so a file that is locked
        # by Excel does not crash the reader.
        suffix = Path(file_path).suffix
        if suffix.lower() not in (".xlsx", ".xlsm", ".xltx"):
            raise DataReaderError(
                f"Unsupported file type '{suffix}'. Only .xlsx files are "
                "supported (openpyxl cannot read legacy .xls)."
            )
        temp_handle = tempfile.NamedTemporaryFile(
            suffix=suffix, delete=False
        )
        temp_path = temp_handle.name
        temp_handle.close()
        try:
            shutil.copyfile(file_path, temp_path)
            workbook = load_workbook(
                temp_path, data_only=True, read_only=True
            )

            if self.config.get("auto"):
                detection = AutoDetector(self.logger).detect(workbook)
                if detection is None:
                    raise DataReaderError(
                        "Could not auto-detect any data in the file."
                    )
                self._log(
                    "Auto-detected: "
                    f"sheet='{detection.worksheet_name}', "
                    f"layout={detection.layout}, "
                    f"date_col={detection.date_col}, "
                    f"metric_col={detection.metric_col}, "
                    f"value_col={detection.value_col}, "
                    f"date_format={detection.date_format}"
                )
                worksheet = workbook[detection.worksheet_name]
                records = self._read_auto(worksheet, detection)
            else:
                worksheet = self._pick_worksheet(workbook)
                records = self._read_configured(worksheet)

            workbook.close()
            records.sort(key=lambda r: r["timestamp"])
            self._log(f"Read {len(records)} clean records from {file_path}")
            return records
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
